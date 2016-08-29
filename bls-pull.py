import ConfigParser
import requests
import json
from openpyxl import Workbook

#API Key stored locally in config file
parser = ConfigParser.ConfigParser()
#Change the .ini file to where your API key is stored
parser.read("C:\\apikeys.ini")
#APIkeys is section header, FRED is variable.
#See https://roboticape.wordpress.com/2014/01/03/hiding-secrets-from-github-or-using-configparser/ for method
key = parser.get('APIkeys', 'BLS')

#BLS series names and series IDs to pull. Change the names of these variables and the series numbers to pull different data.
DCTotalNonFarmEmployment = 'SMS11000000000000001'
DCTotalPrivateEmployment = 'SMS11000000500000001'

#Enter the start and end years for the data you want to pull
startyear = '2000'
endyear = '2016'

#Enter the filename for the excel file you will save
filename = 'Total-Private-Employment-test'

# API pull data below taken from sample at http://www.bls.gov/developers/api_python.htm#python2
headers = {'Content-type': 'application/json'}

data = json.dumps({"seriesid": [DCTotalNonFarmEmployment, DCTotalPrivateEmployment], "startyear":startyear, "endyear":endyear, 'registrationKey':key})
p = requests.post('http://api.bls.gov/publicAPI/v2/timeseries/data/', data=data, headers=headers)

json_data = json.loads(p.text)

#Debugging code for checking that the data is scraping properly
#for series in json_data['Results']['series']:
	# Data is a list of dicts
	#for record in series['data']:
	#	print series['seriesID']
	#	print record['year']
	#	print record['periodName']
	#	print record['value']
	#	print "\n"

# Create a workbook with a sheet named ForTableau
wb = Workbook(guess_types=True)
ForTableau = wb.active
ForTableau.title = "ForTableau"

# Write the column headers
ForTableau['A1'] = 'Date'
ForTableau['B1'] = 'Total Nonfarm Employment'
ForTableau['C1'] = 'Total Private Employment'

for series in json_data['Results']['series']:
	if series['seriesID'] == DCTotalNonFarmEmployment:
		for i, record in enumerate(series['data']):
			a = ForTableau.cell(row = i+2, column = 1)
			a.value = record['periodName']+'-01-'+record['year']
			b = ForTableau.cell(row = i+2, column = 2)
			b.value = float(record['value'])*1000

	elif series['seriesID'] == DCTotalPrivateEmployment:
		for i, record in enumerate(series['data']):
			c= ForTableau.cell(row = i+2, column = 3)
			c.value = float(record['value'])*1000

wb.save(filename+'.xlsx')
