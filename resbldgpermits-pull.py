import requests
import mechanize
from openpyxl import Workbook
from bs4 import BeautifulSoup
from datetime import datetime

# Create a workbook with a sheet named ForTableau
wb = Workbook(guess_types=True)
ForTableau = wb.create_sheet()
ForTableau.title = "ForTableau"

# Create the empty list that will fill with lists of data for each month/bldgtype.
monthdata = []

# Write the column headers
ForTableau['A1'] = 'Building Type'
ForTableau['B1'] = 'Date'
ForTableau['C1'] = 'Buildings'
ForTableau['D1'] = 'Units'
ForTableau['E1'] = 'Construction Cost'
ForTableau['F1'] = 'Moving Average'

# Emulate a browser with mechanize
br = mechanize.Browser()

# Browser options
br.set_handle_equiv(True)
br.set_handle_referer(True)
br.set_handle_robots(False)
br.addheaders = [('User-agent', 'XX' )] # ADD YOUR HEADERS

# Create list with month numbers to fill in form.
months = []
for x in range(1,13):
	months.append(str(x).zfill(2))

# Create list with years to fill in form.
years = []
for x in range(2005,2016): # CHANGE THE YEARS HERE IF NEEDED
	years.append(str(x))

# Loop over month/year
for year in years:
	for month in months:
		br.open('http://censtats.census.gov/bldg/bldgprmt.shtml')

		# To fill in the first page's form with DC and data format
		for form in br.forms():
		br.select_form("bldgform")

		br.form['o'] = ['Monthly',]
		br.form['M'] = [month]
		br.form['S'] = ['11Distrirct of Columbia'] # Yes, Census misspelled 'District'

		br.submit()

		#Select the right form on the second page.
		br.select_form(nr=1)
		#Make the controls in the form non-read only; we need to back-fill some values.
		br.form.set_all_readonly(False)

		#This edits the year selection to be 2015. 
		br.form['Y'] = year

		#The control for DC's "Counties" is empty, so create the right object in it. 
		item = mechanize.Item(br.form.find_control(name='C'), {'contents': '001000', 'value': '001000', 'label': '001000'})
		br.form['C'] = ['001000']

		# Store the html of the page to parse. 
		response2 = br.submit()
		site3 = response2.read()

		# The data we want is uniquely in a table with a border of 1	
		soup = BeautifulSoup(site3, 'html.parser')
		table = soup.find("table", border=1)

		# Scrape the relevant figures from the table and store as a record. 
		try:
			for each in table.findAll('tr')[4:8]:
				col = each.findAll('td')
				bldgtype = col[1].string
				date = month + "-01-" + year
				buildings = col[2].string
				units = col[3].string
				constructioncost = col[4].string
				record = (bldgtype, date, buildings, units, constructioncost)
				monthdata.append(record)
		# Print out which month-years failed to produce data. 
		except AttributeError: 
			print (year + "-" + month + " not ready yet.")

numrecords = len(monthdata)
a=int(0)
b=int(0)

for i in range(2, (numrecords+2)):
	for j in range(1,6):
		c = ForTableau.cell(row = i, column = j)
		c.value = monthdata[a][j-1]
	a += 1

for i in range(10, (numrecords+2)):
	d = ForTableau.cell(row = i, column = 6)
	x = ForTableau.cell(row = (i - 8), column = 4)
	y = ForTableau.cell(row = (i - 4), column = 4)
	z = ForTableau.cell(row = i, column = 4)
	avg = ((x.value + y.value + z.value)/3.00)
	d.value = avg
	d.number_format = '0.00'

wb.save('DC-Housing Units Authorized.xlsx')
